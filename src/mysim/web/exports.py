"""Export utilities for CSV and XLSX generation."""

from __future__ import annotations

import csv
import io
import time
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, numbers
from openpyxl.utils import get_column_letter


def generate_csv(results: list[dict[str, Any]], scenario_name: str) -> tuple[str, str]:
    """Generate a CSV export with German locale (semicolon delimiter).
    
    Returns (filename, csv_content).
    """
    if not results:
        return f"{scenario_name}_empty.csv", ""

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{scenario_name}_{timestamp}.csv"

    output = io.StringIO()
    # Exclude traces from CSV
    fieldnames = [k for k in results[0].keys() if k != "traces"]

    writer = csv.DictWriter(
        output, fieldnames=fieldnames, delimiter=";", extrasaction="ignore"
    )
    writer.writeheader()

    for row in results:
        formatted_row = {}
        for key in fieldnames:
            val = row.get(key)
            if isinstance(val, Decimal):
                # German decimal format for CSV
                formatted_row[key] = str(val).replace(".", ",")
            elif isinstance(val, bool):
                formatted_row[key] = "Ja" if val else "Nein"
            else:
                formatted_row[key] = val
        writer.writerow(formatted_row)

    return filename, output.getvalue()


def generate_xlsx(
    results: list[dict[str, Any]],
    scenario_name: str,
    include_traces: bool = False,
) -> tuple[str, bytes]:
    """Generate an XLSX export with proper formatting.
    
    Returns (filename, xlsx_bytes).
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{scenario_name}_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Simulation"

    if not results:
        wb_bytes = _workbook_to_bytes(wb)
        return filename, wb_bytes

    # Determine columns (exclude traces unless requested)
    all_keys = list(results[0].keys())
    if not include_traces:
        all_keys = [k for k in all_keys if k != "traces"]

    # Write headers with bold font
    bold_font = Font(bold=True)
    for col_idx, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = bold_font

    # Write data rows
    for row_idx, row in enumerate(results, 2):
        for col_idx, key in enumerate(all_keys, 1):
            val = row.get(key)
            if isinstance(val, Decimal):
                cell = ws.cell(row=row_idx, column=col_idx, value=float(val))
                cell.number_format = '#.##0,00'
            elif isinstance(val, bool):
                ws.cell(row=row_idx, column=col_idx, value="Ja" if val else "Nein")
            elif isinstance(val, (list, dict)):
                ws.cell(row=row_idx, column=col_idx, value=str(val))
            else:
                ws.cell(row=row_idx, column=col_idx, value=val)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-width columns
    for col_idx, key in enumerate(all_keys, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(len(key) + 2, 14)

    xlsx_bytes = _workbook_to_bytes(wb)
    return filename, xlsx_bytes


def _workbook_to_bytes(wb: Workbook) -> bytes:
    """Serialize workbook to bytes."""
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def cleanup_old_exports(export_dir: Path, max_age_seconds: int = 3600) -> None:
    """Delete export files older than max_age_seconds."""
    if not export_dir.exists():
        return
    now = time.time()
    for f in export_dir.iterdir():
        if f.is_file() and (now - f.stat().st_mtime) > max_age_seconds:
            f.unlink(missing_ok=True)
