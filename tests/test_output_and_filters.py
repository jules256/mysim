from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from mysim.output import results_to_json, results_to_table
from mysim.web.exports import generate_csv, generate_xlsx
from mysim.web.filters import format_german_currency, format_german_number, register_filters
from mysim.web.app import create_app


def test_results_to_json_serializes_decimal():
    results = [{"year": 2026, "value": Decimal("123.45")}]
    json_text = results_to_json(results)
    assert '"123.45"' in json_text
    assert '"year"' in json_text


def test_results_to_table_contains_rows():
    results = [{
        "year": 2026,
        "age": 52,
        "total_inflows": Decimal("1000"),
        "total_outflows": Decimal("500"),
        "total_deductions": Decimal("100"),
        "net_annual_result": Decimal("400"),
        "capital_total_savings": Decimal("1000"),
    }]
    table = results_to_table(results)
    assert "Jahr" in table
    assert "1,000.00" in table


def test_generate_csv_and_xlsx_exports():
    results = [{
        "year": 2026,
        "age": 52,
        "total_inflows": Decimal("1000"),
        "total_outflows": Decimal("500"),
        "total_deductions": Decimal("100"),
        "net_annual_result": Decimal("400"),
        "capital_total_savings": Decimal("1000"),
        "is_insolvent": False,
    }]

    filename, csv_content = generate_csv(results, "scenario")
    assert filename.startswith("scenario_")
    assert ";" in csv_content
    assert "1000;" in csv_content

    filename, xlsx_bytes = generate_xlsx(results, "scenario")
    workbook = load_workbook(filename=BytesIO(xlsx_bytes))
    sheet = workbook.active
    assert sheet.cell(row=1, column=1).value == "year"
    assert sheet.cell(row=2, column=1).value == 2026


def test_german_filters_format_currency_and_numbers():
    assert format_german_currency(Decimal("1234.56")) == "1.234,56 €"
    assert format_german_currency("-1000") == "-1.000,00 €"
    assert format_german_currency("hello") == "hello"
    assert format_german_currency(None) == "—"
    assert format_german_number(1234567) == "1.234.567"
    assert format_german_number(True) == "Ja"
    assert format_german_number("not a number") == "not a number"
    assert format_german_number(None) == "—"


def test_register_filters_attaches_jinja_filters():
    app = create_app(scenarios_dir="./scenarios")
    register_filters(app)
    assert app.jinja_env.filters["german_currency"] is not None
    assert app.jinja_env.filters["german_number"] is not None
